from fastapi import FastAPI, Depends, HTTPException, status, Request, Response as FastAPIResponse, Header, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.security import HTTPBearer, HTTPBasic, HTTPBasicCredentials
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from pydantic import BaseModel, EmailStr
from datetime import timedelta, datetime
from typing import List, Optional
from contextlib import asynccontextmanager
import uvicorn
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import mysql.connector
import random
import string
import os
import secrets
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
import json
from typing import Dict, Set

from database import get_db, create_tables, Admin, Event, Student, User, EmailSettings, EmailTemplate, Form, FormQuestion, FormResponse, FormAnalytics, QAQuestion, UserQuestionCount, OTP
from auth import verify_password, get_password_hash, create_access_token, verify_token, ACCESS_TOKEN_EXPIRE_MINUTES
from chat_routes import router as chat_router
from chat_models import Question, Poll, PollResult, ChatAdmin
from forms_routes import router as forms_router


from payment_api import router as payment_router
from payment_model import Payment

@asynccontextmanager
async def lifespan(app: FastAPI):
    # Startup
    create_tables()

    yield
    # Shutdown (if needed)

app = FastAPI(title="Dashboard API", version="1.0.0", lifespan=lifespan)

# WebSocket connection manager for forms
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, Set[WebSocket]] = {}
    
    async def connect(self, websocket: WebSocket, form_id: str):
        await websocket.accept()
        if form_id not in self.active_connections:
            self.active_connections[form_id] = set()
        self.active_connections[form_id].add(websocket)
    
    def disconnect(self, websocket: WebSocket, form_id: str):
        if form_id in self.active_connections:
            self.active_connections[form_id].discard(websocket)
            if not self.active_connections[form_id]:
                del self.active_connections[form_id]
    
    async def send_personal_message(self, message: str, websocket: WebSocket):
        await websocket.send_text(message)
    
    async def broadcast_to_form(self, message: str, form_id: str):
        if form_id in self.active_connections:
            for connection in self.active_connections[form_id].copy():
                try:
                    await connection.send_text(message)
                except:
                    self.active_connections[form_id].discard(connection)

manager = ConnectionManager()

# Include chat routes
app.include_router(chat_router, prefix="/api")

# WebSocket manager for Q/A system
class QAConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, WebSocket] = {}
    
    async def connect(self, websocket: WebSocket, user_email: str):
        await websocket.accept()
        self.active_connections[user_email] = websocket
    
    def disconnect(self, user_email: str):
        if user_email in self.active_connections:
            del self.active_connections[user_email]
    
    async def send_personal_message(self, message: str, user_email: str):
        if user_email in self.active_connections:
            try:
                await self.active_connections[user_email].send_text(message)
            except:
                self.disconnect(user_email)
    
    async def broadcast(self, message: str):
        for user_email, connection in list(self.active_connections.items()):
            try:
                await connection.send_text(message)
            except:
                self.disconnect(user_email)

qa_manager = QAConnectionManager()

# Include forms routes
app.include_router(forms_router, prefix="/api")

# Include payment routes
app.include_router(payment_router, prefix="/api")

# Mount static files for serving uploaded images
os.makedirs("uploads", exist_ok=True)
app.mount("/api/files", StaticFiles(directory="uploads"), name="files")



# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "DELETE", "OPTIONS"],
    allow_headers=["*"],
)

# Pydantic models
class LoginRequest(BaseModel):
    email: str
    password: str

class LoginResponse(BaseModel):
    access_token: str
    token_type: str
    user: dict

class AdminCreate(BaseModel):
    email: str
    role: str
    password: str

class PasswordResetRequest(BaseModel):
    new_password: str
    confirm_password: str

class QAValidationRequest(BaseModel):
    email: str
    registration_id: str

class QAQuestionRequest(BaseModel):
    user_email: str
    user_name: str
    registration_id: str
    question: str

class QAToggleRequest(BaseModel):
    event_id: int
    active: bool

class QAActionRequest(BaseModel):
    question_id: int
    action: str  # answered, skipped, rejected
    response: str = ""

class EventCreate(BaseModel):
    name: str
    slug: str
    event_date: str
    event_time: str
    
    def validate_fields(self):
        if not self.name or not self.name.strip():
            raise ValueError("Event name is required")
        if not self.slug or not self.slug.strip():
            raise ValueError("Event slug is required")
        if not self.event_date or not self.event_date.strip():
            raise ValueError("Event date is required")
        if not self.event_time or not self.event_time.strip():
            raise ValueError("Event time is required")
        return True

class EventUpdate(BaseModel):
    name: str
    slug: str
    event_date: str
    event_time: str
    
    def validate_fields(self):
        if not self.name or not self.name.strip():
            raise ValueError("Event name is required")
        if not self.slug or not self.slug.strip():
            raise ValueError("Event slug is required")
        if not self.event_date or not self.event_date.strip():
            raise ValueError("Event date is required")
        if not self.event_time or not self.event_time.strip():
            raise ValueError("Event time is required")
        return True

class UserCreate(BaseModel):
    name: str
    first_name: str
    last_name: str
    email: str
    phone_number: str
    college_name: str
    eventId: int

class EmailTemplateCreate(BaseModel):
    name: str
    subject: str
    content: str

class EmailSettingsCreate(BaseModel):
    smtp_server: str
    smtp_port: int
    smtp_username: str
    smtp_password: str
    from_email: str

class SendEmailRequest(BaseModel):
    template_id: int
    student_ids: List[str]

class TestEmailRequest(BaseModel):
    email: str

# Workshop Registration Models
class CustomFields(BaseModel):
    primaryEmail: Optional[str] = None
    collegeName: Optional[str] = None
    yearSemester: Optional[str] = None
    course: Optional[str] = None
    specifyCourse: Optional[str] = None
    howDidYouHear: Optional[str] = None
    referralEmail: Optional[str] = None
    userType: str
    isCurrentStudent: Optional[str] = None
    registrationId: Optional[str] = None
    gender: Optional[str] = None
    agreeToTerms: str

class UserRegistration(BaseModel):
    name: str
    firstName: str
    lastName: Optional[str] = None
    email: EmailStr
    phone_number: str
    custom_fields: CustomFields
    project: str = "kambaa.ai"
    formName: str = "AI-workshop-new"
    eventId: str
    
    def validate_emails(self):
        """Validate that referralEmail is different from email and primaryEmail"""
        main_email = str(self.email).lower()
        
        if self.custom_fields.referralEmail and self.custom_fields.referralEmail.strip():
            referral_email = self.custom_fields.referralEmail.lower()
            
            if referral_email == main_email:
                raise ValueError("Referral Email must be different from main Email")
            
            if (self.custom_fields.primaryEmail and 
                self.custom_fields.primaryEmail.strip() and 
                referral_email == self.custom_fields.primaryEmail.lower()):
                raise ValueError("Referral Email must be different from Primary Email")
        
        return True

class OTPRequest(BaseModel):
    email: EmailStr

class OTPVerification(BaseModel):
    email: EmailStr
    otp: str

class Response(BaseModel):
    message: str
    success: bool

# Static token for workshop registration
STATIC_TOKEN = "kambaa-ai-workshop-2024-static-token"

def verify_static_token(authorization: str = Header(None)):
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Missing or invalid authorization header")
    
    token = authorization.split(" ")[1]
    if token != STATIC_TOKEN:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    return token



# Q/A Session endpoints
@app.get("/api/qa/active-event")
def get_active_qa_event(db: Session = Depends(get_db)):
    try:
        print("Checking for active Q/A event...")
        
        # Check if there's an active event with Q/A enabled
        active_event = db.query(Event).filter(Event.qa_active == 1).first()
        if not active_event:
            print("No active Q/A event found")
            # List all events for debugging
            all_events = db.query(Event).all()
            print(f"All events: {[(e.id, e.name, getattr(e, 'qa_active', 0)) for e in all_events]}")
            raise HTTPException(status_code=404, detail="No active Q/A session found")
        
        print(f"Active Q/A event found: {active_event.name} (ID: {active_event.id})")
        
        return {
            "id": active_event.id,
            "name": active_event.name,
            "event_date": str(active_event.event_date) if active_event.event_date else "",
            "qa_active": active_event.qa_active,
            "status": "active"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in get_active_qa_event: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/api/qa/validate-user")
def validate_qa_user(validation_data: QAValidationRequest, db: Session = Depends(get_db)):
    try:
        print(f"Validating user: email={validation_data.email}, reg_id={validation_data.registration_id}")
        
        # Get the active event first
        active_event = db.query(Event).filter(Event.qa_active == 1).first()
        if not active_event:
            print("No active Q/A event found")
            raise HTTPException(status_code=404, detail="No active Q/A session found")
        
        print(f"Active Q/A event: {active_event.name} (ID: {active_event.id})")
        
        # Find user by email and registration_id
        user = db.query(User).filter(
            User.email == validation_data.email,
            User.registration_id == validation_data.registration_id
        ).first()
        
        if not user:
            print(f"User not found with email: {validation_data.email} and reg_id: {validation_data.registration_id}")
            # Debug: Check if user exists with just email
            user_by_email = db.query(User).filter(User.email == validation_data.email).first()
            if user_by_email:
                print(f"User exists with email but different reg_id. Expected: {validation_data.registration_id}, Found: {user_by_email.registration_id}")
            raise HTTPException(status_code=404, detail="Invalid credentials. Please check your registration ID.")
        
        print(f"User found: {user.name} (Email: {user.email})")
        
        # Check if user is registered for the active event
        if user.eventId != active_event.id:
            print(f"User event ID ({user.eventId}) doesn't match active event ID ({active_event.id})")
            # Still allow if user is registered for any event
            pass
        
        user_name = user.name or f"{user.first_name or ''} {user.last_name or ''}".strip()
        
        return {
            "user_name": user_name,
            "user_email": user.email,
            "event_id": active_event.id,
            "event_name": active_event.name,
            "valid": True
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in validate_qa_user: {str(e)}")
        raise HTTPException(status_code=500, detail="Internal server error")

@app.post("/api/qa/submit-question")
def submit_qa_question(question_data: QAQuestionRequest, db: Session = Depends(get_db)):
    try:
        print(f"Submitting question from user: {question_data.user_email}")
        
        # Get active event first
        active_event = db.query(Event).filter(Event.qa_active == 1).first()
        if not active_event:
            print("No active Q/A session")
            raise HTTPException(status_code=404, detail="No active Q/A session")
        
        # Verify user exists
        user = db.query(User).filter(
            User.email == question_data.user_email,
            User.registration_id == question_data.registration_id
        ).first()
        
        if not user:
            print(f"User not found: {question_data.user_email} with reg_id: {question_data.registration_id}")
            raise HTTPException(status_code=404, detail="User not found")
        
        # Create new question
        new_question = QAQuestion(
            user_email=question_data.user_email,
            user_name=question_data.user_name,
            registration_id=question_data.registration_id,
            question=question_data.question.strip(),
            event_id=active_event.id,
            status="pending"
        )
        
        db.add(new_question)
        db.commit()
        db.refresh(new_question)
        
        print(f"Question submitted successfully with ID: {new_question.id}")
        
        return {
            "message": "Question submitted successfully", 
            "question_id": new_question.id,
            "status": "success"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error submitting question: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail="Failed to submit question")

@app.post("/api/qa/toggle")
def toggle_qa_session(toggle_data: QAToggleRequest, db: Session = Depends(get_db)):
    # First, disable all other Q/A sessions
    db.query(Event).update({Event.qa_active: 0})
    
    # Get the specific event
    event = db.query(Event).filter(Event.id == toggle_data.event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")
    
    # Set the Q/A status for this event
    event.qa_active = 1 if toggle_data.active else 0
    db.commit()
    
    status = "enabled" if toggle_data.active else "disabled"
    return {"message": f"Q/A session {status} for event: {event.name}"}

@app.post("/api/auth/login", response_model=LoginResponse)
def login(login_data: LoginRequest, db: Session = Depends(get_db)):
    print(f"Login attempt for email: {login_data.email}")
    
    # Find admin by email
    admin = db.query(Admin).filter(Admin.email == login_data.email).first()
    print(f"Admin found: {admin is not None}")
    
    if not admin or not verify_password(login_data.password, admin.hashed_password):
        print("Login failed: Invalid credentials")
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect email or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    print("Login successful")
    
    # Create access token
    access_token_expires = timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = create_access_token(
        data={"sub": admin.email, "role": admin.role}, expires_delta=access_token_expires
    )
    
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": admin.id,
            "email": admin.email,
            "role": admin.role
        }
    }

@app.get("/api/auth/me")
def get_current_user(current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not admin:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {
        "id": admin.id,
        "email": admin.email,
        "role": admin.role
    }

@app.post("/api/auth/create-admin")
def create_admin(admin_data: AdminCreate, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Check if current user is admin or manager (presenters cannot create users)
    current_admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not current_admin or current_admin.role not in ["admin", "manager"]:
        raise HTTPException(status_code=403, detail="Only admins and managers can create new users")
    
    # Validate role
    if admin_data.role not in ["admin", "manager", "presenter"]:
        raise HTTPException(status_code=400, detail="Invalid role. Must be admin, manager, or presenter")
    
    existing_admin = db.query(Admin).filter(Admin.email == admin_data.email).first()
    
    if existing_admin:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Admin with this email already exists"
        )
    
    hashed_password = get_password_hash(admin_data.password)
    new_admin = Admin(
        email=admin_data.email,
        role=admin_data.role,
        hashed_password=hashed_password
    )
    
    db.add(new_admin)
    db.commit()
    db.refresh(new_admin)
    
    return {"message": "Admin created successfully", "admin_id": new_admin.id}

@app.post("/api/auth/reset-password")
def reset_password(password_data: PasswordResetRequest, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Validate passwords match
    if password_data.new_password != password_data.confirm_password:
        raise HTTPException(status_code=400, detail="Passwords do not match")
    
    # Find current user
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not admin:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Update password
    admin.hashed_password = get_password_hash(password_data.new_password)
    db.commit()
    
    return {"message": "Password updated successfully"}

class EmailValidationRequest(BaseModel):
    email: EmailStr
    registration_id: str

@app.post("/api/auth/validate-email")
def validate_email(validation_data: EmailValidationRequest, db: Session = Depends(get_db)):
    """Validate user email and registration ID for form authentication"""
    try:
        # Find user by email and registration_id
        user = db.query(User).filter(
            User.email == validation_data.email,
            User.registration_id == validation_data.registration_id
        ).first()
        
        if not user:
            raise HTTPException(status_code=404, detail="Invalid email or registration ID")
        
        user_name = user.name or f"{user.first_name or ''} {user.last_name or ''}".strip()
        
        return {
            "valid": True,
            "user_name": user_name,
            "user_email": user.email,
            "registration_id": user.registration_id,
            "message": "User validated successfully"
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error validating email: {str(e)}")
        raise HTTPException(status_code=500, detail="Validation failed")

@app.get("/api/events")
def get_events(current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    events = db.query(Event).all()
    result = []
    for e in events:
        # Parse event date and time properly
        event_date_str = str(e.event_date) if e.event_date else ""
        
        # Try to parse the date string and format it properly
        try:
            if event_date_str and event_date_str != "None":
                # Handle different possible formats
                if ' ' in event_date_str:
                    # Format: "2024-01-15 14:30:00" or "2024-01-15 14:30"
                    parts = event_date_str.split(' ', 1)
                    date_part = parts[0]
                    time_part = parts[1] if len(parts) > 1 else "00:00"
                else:
                    # Only date provided
                    date_part = event_date_str
                    time_part = "00:00"
                
                # Validate and format the datetime
                from datetime import datetime
                try:
                    # Parse to validate
                    dt = datetime.strptime(f"{date_part} {time_part}", "%Y-%m-%d %H:%M:%S" if ":" in time_part and time_part.count(":") == 2 else "%Y-%m-%d %H:%M")
                    formatted_datetime = dt.strftime("%Y-%m-%d %H:%M:%S")
                except:
                    # Fallback to original if parsing fails
                    formatted_datetime = event_date_str
                    date_part = event_date_str.split(' ')[0] if ' ' in event_date_str else event_date_str
                    time_part = ' '.join(event_date_str.split(' ')[1:]) if ' ' in event_date_str else "00:00"
            else:
                formatted_datetime = ""
                date_part = ""
                time_part = "00:00"
        except Exception as ex:
            # Fallback for any parsing errors
            formatted_datetime = event_date_str
            date_part = event_date_str.split(' ')[0] if event_date_str and ' ' in event_date_str else event_date_str
            time_part = ' '.join(event_date_str.split(' ')[1:]) if event_date_str and ' ' in event_date_str else "00:00"
        
        result.append({
            "id": e.id, 
            "name": e.name, 
            "slug": e.slug or "", 
            "event_date": formatted_datetime,
            "date_part": date_part,
            "time_part": time_part,
            "qa_active": e.qa_active or 0, 
            "created_at": str(e.created_at)
        })
    return result

@app.post("/api/events")
def create_event(event_data: EventCreate, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    try:
        # Check if user is presenter (view-only)
        admin = db.query(Admin).filter(Admin.email == current_user).first()
        if admin and admin.role == "presenter":
            raise HTTPException(status_code=403, detail="View Only - Presenters cannot create events")
        
        # Validate required fields
        event_data.validate_fields()
        
        # Check if slug already exists
        existing_event = db.query(Event).filter(Event.slug == event_data.slug).first()
        if existing_event:
            raise ValueError("Event slug already exists")
        
        # Combine date and time
        event_datetime = f"{event_data.event_date} {event_data.event_time}"
        
        new_event = Event(name=event_data.name, slug=event_data.slug, event_date=event_datetime)
        db.add(new_event)
        db.commit()
        db.refresh(new_event)
        
        # Log audit action
        admin = db.query(Admin).filter(Admin.email == current_user).first()
        user_role = admin.role if admin else "unknown"
        log_audit_action(db, current_user, user_role, "create_event", "event", new_event.id, f"Created event: {event_data.name}")
        
        return {"id": new_event.id, "name": new_event.name, "slug": new_event.slug, "event_date": str(new_event.event_date)}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error creating event: {str(e)}")

# OTP and Registration endpoints
@app.post("/generate-otp")
def generate_otp(otp_request: OTPRequest, token: str = Depends(verify_static_token), db: Session = Depends(get_db)):
    try:
        # Generate 6-digit OTP
        otp_code = str(random.randint(100000, 999999))
        
        print(f"Generated OTP {otp_code} for email: {otp_request.email}")
        
        # Mark all existing unused OTPs for this email as used (cleanup)
        db.query(OTP).filter(
            OTP.email == str(otp_request.email),
            OTP.is_used == 0
        ).update({OTP.is_used: 1})
        
        # Store new OTP in database with 10 minute expiry
        from datetime import timedelta
        expires_at = datetime.utcnow() + timedelta(minutes=10)
        
        new_otp = OTP(
            email=str(otp_request.email),
            otp_code=otp_code,
            expires_at=expires_at
        )
        db.add(new_otp)
        db.commit()
        
        print(f"Stored OTP in database. Expires at: {expires_at}")
        
        # Send OTP via email
        email_settings = db.query(EmailSettings).first()
        if email_settings:
            try:
                import smtplib
                from email.mime.text import MIMEText
                from email.mime.multipart import MIMEMultipart
                
                server = None
                if email_settings.smtp_port == 465:
                    server = smtplib.SMTP_SSL(email_settings.smtp_server, email_settings.smtp_port)
                else:
                    server = smtplib.SMTP(email_settings.smtp_server, email_settings.smtp_port)
                    server.starttls()
                
                server.login(email_settings.smtp_username, email_settings.smtp_password)
                
                msg = MIMEMultipart()
                msg['From'] = email_settings.from_email
                msg['To'] = str(otp_request.email)
                msg['Subject'] = "Email Verification OTP"
                
                body = f"Your OTP for email verification is: {otp_code}. This OTP will expire in 10 minutes."
                msg.attach(MIMEText(body, 'plain'))
                
                server.send_message(msg)
                server.quit()
                print("OTP email sent successfully")
            except Exception as email_error:
                print(f"Email sending failed: {email_error}")
        
        return {"message": "OTP sent successfully", "success": True}
    except Exception as e:
        print(f"Error in generate_otp: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/verify-otp")
def verify_otp(otp_verification: OTPVerification, token: str = Depends(verify_static_token), db: Session = Depends(get_db)):
    try:
        print(f"Verifying OTP for email: {otp_verification.email}, OTP: {otp_verification.otp}")
        
        # Find the most recent valid OTP for this email and code
        otp_record = db.query(OTP).filter(
            OTP.email == str(otp_verification.email),
            OTP.otp_code == str(otp_verification.otp),
            OTP.is_used == 0
        ).order_by(OTP.created_at.desc()).first()
        
        print(f"Found OTP record: {otp_record is not None}")
        
        if not otp_record:
            print("No OTP record found")
            raise HTTPException(status_code=400, detail="Invalid OTP")
        
        # Check if OTP is expired
        current_time = datetime.utcnow()
        if otp_record.expires_at <= current_time:
            print(f"OTP expired. Current: {current_time}, Expires: {otp_record.expires_at}")
            raise HTTPException(status_code=400, detail="OTP has expired")
        
        print(f"OTP is valid, marking OTP ID {otp_record.id} as used")
        
        # Mark this specific OTP as used using update with ID
        db.query(OTP).filter(OTP.id == otp_record.id).update({OTP.is_used: 1})
        db.commit()
        
        print("OTP marked as used successfully")
        
        return {"message": "OTP verified successfully", "success": True}
    except HTTPException:
        raise
    except Exception as e:
        print(f"Error in verify_otp: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/register")
def register_user(user_data: UserRegistration, token: str = Depends(verify_static_token), db: Session = Depends(get_db)):
    try:
        # Validate emails
        user_data.validate_emails()
        
        # Check if user already exists
        existing_user = db.query(User).filter(User.email == user_data.email).first()
        if existing_user:
            raise HTTPException(status_code=400, detail="User with this email already exists")
        
        # Get event by eventId (slug)
        event = db.query(Event).filter(Event.slug == user_data.eventId).first()
        if not event:
            raise HTTPException(status_code=400, detail="Invalid event")
        
        # Create new user
        new_user = User(
            name=user_data.name,
            first_name=user_data.firstName,
            last_name=user_data.lastName or "",
            email=str(user_data.email),
            phone_number=user_data.phone_number,
            primary_email=user_data.custom_fields.primaryEmail or "",
            college_name=user_data.custom_fields.collegeName or "",
            year_semester=user_data.custom_fields.yearSemester or "",
            course=user_data.custom_fields.course or "",
            specify_course=user_data.custom_fields.specifyCourse or "",
            how_did_you_hear=user_data.custom_fields.howDidYouHear or "",
            referral_email=user_data.custom_fields.referralEmail or "",
            user_type=user_data.custom_fields.userType,
            is_current_student=user_data.custom_fields.isCurrentStudent or "",
            registration_id=user_data.custom_fields.registrationId or "",
            gender=user_data.custom_fields.gender or "",
            agree_to_terms=user_data.custom_fields.agreeToTerms,
            project=user_data.project,
            form_name=user_data.formName,
            eventId=event.id,
            email_verified=1
        )
        
        db.add(new_user)
        db.commit()
        db.refresh(new_user)
        
        return {
            "message": "Registration successful",
            "success": True,
            "user_id": new_user.id,
            "registration_id": new_user.registration_id
        }
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/api/events/{event_id}")
def update_event(event_id: int, event_data: EventUpdate, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    try:
        # Check if user is presenter (view-only)
        admin = db.query(Admin).filter(Admin.email == current_user).first()
        if admin and admin.role == "presenter":
            raise HTTPException(status_code=403, detail="View Only - Presenters cannot edit events")
        
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        # Validate required fields
        event_data.validate_fields()
        
        # Check if slug already exists (excluding current event)
        existing_event = db.query(Event).filter(Event.slug == event_data.slug, Event.id != event_id).first()
        if existing_event:
            raise ValueError("Event slug already exists")
        
        # Combine date and time
        event_datetime = f"{event_data.event_date} {event_data.event_time}"
        
        event.name = event_data.name
        event.slug = event_data.slug
        event.event_date = event_datetime
        db.commit()
        
        # Log audit action
        admin = db.query(Admin).filter(Admin.email == current_user).first()
        user_role = admin.role if admin else "unknown"
        log_audit_action(db, current_user, user_role, "edit_event", "event", event_id, f"Updated event: {event_data.name}")
        
        return {"id": event.id, "name": event.name, "slug": event.slug, "event_date": str(event.event_date)}
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error updating event: {str(e)}")

@app.delete("/api/events/{event_id}")
def delete_event(event_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    try:
        # Check if user is presenter (view-only)
        admin = db.query(Admin).filter(Admin.email == current_user).first()
        if admin and admin.role == "presenter":
            raise HTTPException(status_code=403, detail="View Only - Presenters cannot delete events")
        
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            raise HTTPException(status_code=404, detail="Event not found")
        
        event_name = event.name
        db.delete(event)
        db.commit()
        
        # Log audit action
        admin = db.query(Admin).filter(Admin.email == current_user).first()
        user_role = admin.role if admin else "unknown"
        log_audit_action(db, current_user, user_role, "delete_event", "event", event_id, f"Deleted event: {event_name}")
        
        return {"message": "Event deleted successfully"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Error deleting event: {str(e)}")

@app.get("/api/events/{event_id}/students")
def get_event_students(event_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    users = db.query(User).filter(User.eventId == event_id).all()
    return [{
        "id": u.id, 
        "name": u.name or f"{u.first_name or ''} {u.last_name or ''}".strip(), 
        "email": u.email, 
        "phone": u.phone_number, 
        "college": u.college_name,
        "registered_at": str(u.created_at)
    } for u in users]

@app.post("/api/users")
def create_user(user_data: UserCreate, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    new_user = User(
        name=user_data.name,
        first_name=user_data.first_name,
        last_name=user_data.last_name,
        email=user_data.email,
        phone_number=user_data.phone_number,
        college_name=user_data.college_name,
        eventId=user_data.eventId
    )
    db.add(new_user)
    db.commit()
    db.refresh(new_user)
    return {"id": new_user.id, "name": new_user.name, "email": new_user.email}

# Email Templates endpoints - using MySQL
@app.get("/api/email-templates")
def get_email_templates(current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    templates = db.query(EmailTemplate).all()
    return [{
        "id": t.id,
        "name": t.name,
        "subject": t.subject,
        "content": t.content,
        "created_at": str(t.created_at)
    } for t in templates]

@app.post("/api/email-templates")
def create_email_template(template_data: EmailTemplateCreate, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    new_template = EmailTemplate(
        name=template_data.name,
        subject=template_data.subject,
        content=template_data.content,
        created_by=current_user
    )
    db.add(new_template)
    db.commit()
    db.refresh(new_template)
    
    # Log audit action
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    user_role = admin.role if admin else "unknown"
    log_audit_action(db, current_user, user_role, "create_email_template", "email_template", new_template.id, f"Created email template: {template_data.name}")
    
    return {"id": new_template.id, "message": "Template created successfully"}

@app.get("/api/email-templates/{template_id}")
def get_email_template(template_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    template = db.query(EmailTemplate).filter(EmailTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    return {
        "id": template.id,
        "name": template.name,
        "subject": template.subject,
        "content": template.content,
        "created_at": str(template.created_at),
        "created_by": template.created_by
    }

@app.put("/api/email-templates/{template_id}")
def update_email_template(template_id: int, template_data: EmailTemplateCreate, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    template = db.query(EmailTemplate).filter(EmailTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    template.name = template_data.name
    template.subject = template_data.subject
    template.content = template_data.content
    template.updated_at = datetime.utcnow()
    
    db.commit()
    
    # Log audit action
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    user_role = admin.role if admin else "unknown"
    log_audit_action(db, current_user, user_role, "edit_email_template", "email_template", template_id, f"Updated email template: {template_data.name}")
    
    return {"id": template.id, "message": "Template updated successfully"}

@app.delete("/api/email-templates/{template_id}")
def delete_email_template(template_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    template = db.query(EmailTemplate).filter(EmailTemplate.id == template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Log audit action before deletion
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    user_role = admin.role if admin else "unknown"
    template_name = template.name
    log_audit_action(db, current_user, user_role, "delete_email_template", "email_template", template_id, f"Deleted email template: {template_name}")
    
    db.delete(template)
    db.commit()
    return {"message": "Template deleted successfully"}

# Email Settings endpoints
@app.get("/api/email-settings")
def get_email_settings(current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    settings = db.query(EmailSettings).first()
    if not settings:
        return None
    return {
        "id": settings.id,
        "smtp_server": settings.smtp_server,
        "smtp_port": settings.smtp_port,
        "smtp_username": settings.smtp_username,
        "smtp_password": "••••••••••••••••" if settings.smtp_password else "",
        "from_email": settings.from_email
    }

@app.post("/api/email-settings")
def save_email_settings(settings_data: EmailSettingsCreate, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    existing_settings = db.query(EmailSettings).first()
    if existing_settings:
        existing_settings.smtp_server = settings_data.smtp_server
        existing_settings.smtp_port = settings_data.smtp_port
        existing_settings.smtp_username = settings_data.smtp_username
        if settings_data.smtp_password and settings_data.smtp_password != "••••••••••••••••":
            existing_settings.smtp_password = settings_data.smtp_password
        existing_settings.from_email = settings_data.from_email
        existing_settings.updated_at = datetime.utcnow()
    else:
        new_settings = EmailSettings(
            smtp_server=settings_data.smtp_server,
            smtp_port=settings_data.smtp_port,
            smtp_username=settings_data.smtp_username,
            smtp_password=settings_data.smtp_password,
            from_email=settings_data.from_email
        )
        db.add(new_settings)
    
    db.commit()
    return {"message": "Email settings saved successfully"}

# Test email connection
@app.post("/api/test-email-connection")
def test_email_connection(test_request: TestEmailRequest, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Get email settings
    email_settings = db.query(EmailSettings).first()
    if not email_settings:
        raise HTTPException(status_code=400, detail="Email settings not configured")
    
    # Validate required fields
    if not email_settings.smtp_server or not email_settings.smtp_username or not email_settings.smtp_password:
        raise HTTPException(status_code=400, detail="Incomplete email settings. Please configure all SMTP fields.")
    
    server = None
    try:
        # Try SSL first (port 465), then STARTTLS (port 587)
        if email_settings.smtp_port == 465:
            # Use SSL
            server = smtplib.SMTP_SSL(email_settings.smtp_server, email_settings.smtp_port, timeout=30)
        else:
            # Use STARTTLS
            server = smtplib.SMTP(email_settings.smtp_server, email_settings.smtp_port, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()
        
        # Login
        server.login(email_settings.smtp_username, email_settings.smtp_password)
        
        # Send test email
        msg = MIMEMultipart()
        msg['From'] = email_settings.from_email or email_settings.smtp_username
        msg['To'] = test_request.email
        msg['Subject'] = "Test Email Connection"
        
        content = "<h2>Email Connection Test Successful!</h2><p>Your SMTP configuration is working correctly.</p>"
        msg.attach(MIMEText(content, 'html'))
        
        server.send_message(msg)
        
        return {"message": "Test email sent successfully"}
        
    except smtplib.SMTPAuthenticationError as e:
        raise HTTPException(status_code=400, detail="SMTP Authentication failed. Please check your username and password. For Gmail, use an App Password.")
    except smtplib.SMTPConnectError as e:
        raise HTTPException(status_code=400, detail="Cannot connect to SMTP server. Please check server address and port.")
    except smtplib.SMTPServerDisconnected as e:
        raise HTTPException(status_code=400, detail="SMTP server disconnected. Try using port 465 (SSL) or 587 (STARTTLS).")
    except smtplib.SMTPRecipientsRefused as e:
        raise HTTPException(status_code=400, detail="Recipient email address was refused by the server.")
    except Exception as e:
        error_msg = str(e).lower()
        if "getaddrinfo failed" in error_msg or "name or service not known" in error_msg:
            raise HTTPException(status_code=400, detail="Cannot resolve SMTP server address. Please check the server name (e.g., smtp.gmail.com).")
        elif "timed out" in error_msg or "timeout" in error_msg:
            raise HTTPException(status_code=400, detail="Connection timed out. Please check server address and port, or try a different port.")
        elif "ssl" in error_msg or "certificate" in error_msg:
            raise HTTPException(status_code=400, detail="SSL/TLS error. Try using port 587 with STARTTLS or port 465 with SSL.")
        else:
            raise HTTPException(status_code=500, detail=f"Email connection failed: {str(e)}")
    finally:
        if server:
            try:
                server.quit()
            except:
                pass

# Get all users for email sending from MySQL users table only
@app.get("/api/students")
def get_all_students(current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Join users with events to get event names
    users_with_events = db.query(User, Event.name.label('event_name')).outerjoin(
        Event, User.eventId == Event.id
    ).all()
    
    return [{
        "id": str(user.User.id),
        "name": user.User.name or f"{user.User.first_name or ''} {user.User.last_name or ''}".strip(),
        "email": user.User.email,
        "phone": user.User.phone_number,
        "event_id": user.User.eventId,
        "event_name": user.event_name or "Unknown Event",
        "college": user.User.college_name or "",
        "registered_at": str(user.User.created_at)
    } for user in users_with_events]

# Public endpoints
@app.get("/api/public/event/{slug}")
def get_event_by_slug(slug: str, db: Session = Depends(get_db)):
    print(f"Looking for event with slug: {slug}")
    event = db.query(Event).filter(Event.slug == slug).first()
    if not event:
        print(f"Event not found for slug: {slug}")
        # List all events for debugging
        all_events = db.query(Event).all()
        print(f"Available events: {[(e.id, e.slug) for e in all_events]}")
        raise HTTPException(status_code=404, detail="Event not found")
    print(f"Found event: {event.name} with slug: {event.slug}")
    
    # Parse event date and time
    event_date_str = str(event.event_date)
    event_date_parts = event_date_str.split(' ')
    date_part = event_date_parts[0] if len(event_date_parts) > 0 else ''
    time_part = ' '.join(event_date_parts[1:]) if len(event_date_parts) > 1 else ''
    
    return {
        "id": event.id,
        "name": event.name,
        "slug": event.slug,
        "event_date": date_part,
        "event_time": time_part,
        "full_date": event_date_str
    }

@app.get("/api/public/check-user/{email}")
def check_user_exists(email: str, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == email).first()
    return {
        "exists": user is not None,
        "registration_id": getattr(user, 'registration_id', None) if user else None
    }

@app.get("/api/students-protected")
def get_all_students_protected(current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Join users with events to get event names
    users_with_events = db.query(User, Event.name.label('event_name')).outerjoin(
        Event, User.eventId == Event.id
    ).all()
    
    return [{
        "id": str(user.User.id),
        "name": user.User.name or f"{user.User.first_name or ''} {user.User.last_name or ''}".strip(),
        "email": user.User.email,
        "phone": user.User.phone_number,
        "event_id": user.User.eventId,
        "event_name": user.event_name or "Unknown Event",
        "college": user.User.college_name or "",
        "registered_at": str(user.User.created_at)
    } for user in users_with_events]

# Get colleges endpoint
@app.get("/api/colleges")
def get_colleges(current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    from sqlalchemy import func
    colleges = db.query(User.college_name).filter(
        User.college_name.isnot(None),
        User.college_name != ''
    ).distinct().all()
    return {"colleges": [college[0] for college in colleges]}

# Send email endpoint
@app.post("/api/send-email")
def send_email(email_request: SendEmailRequest, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Get email settings
    email_settings = db.query(EmailSettings).first()
    if not email_settings:
        raise HTTPException(status_code=400, detail="Email settings not configured")
    
    # Get email template from MySQL
    template = db.query(EmailTemplate).filter(EmailTemplate.id == email_request.template_id).first()
    if not template:
        raise HTTPException(status_code=404, detail="Template not found")
    
    # Get users from MySQL
    user_ids = [int(id) for id in email_request.student_ids if id.isdigit()]
    users = db.query(User).filter(User.id.in_(user_ids)).all()
    if not users:
        raise HTTPException(status_code=404, detail="No users found")
    
    sent_count = 0
    failed_emails = []
    
    server = None
    try:
        # Setup SMTP with proper SSL/TLS handling
        if email_settings.smtp_port == 465:
            server = smtplib.SMTP_SSL(email_settings.smtp_server, email_settings.smtp_port, timeout=30)
        else:
            server = smtplib.SMTP(email_settings.smtp_server, email_settings.smtp_port, timeout=30)
            server.ehlo()
            server.starttls()
            server.ehlo()
        
        server.login(email_settings.smtp_username, email_settings.smtp_password)
        
        for user in users:
            try:
                # Create personalized email content
                user_name = user.name or f"{user.first_name or ''} {user.last_name or ''}".strip()
                user_email = user.email
                user_phone = user.phone_number or ""
                user_college = user.college_name or ""
                
                # Personalize content
                content = template.content.replace("{{name}}", user_name)
                content = content.replace("{{email}}", user_email)
                content = content.replace("{{phone}}", user_phone)
                content = content.replace("{{college}}", user_college)
                
                # Personalize subject
                subject = template.subject.replace("{{name}}", user_name)
                subject = subject.replace("{{email}}", user_email)
                subject = subject.replace("{{phone}}", user_phone)
                subject = subject.replace("{{college}}", user_college)
                
                msg = MIMEMultipart()
                msg['From'] = email_settings.from_email
                msg['To'] = user_email
                msg['Subject'] = subject
                
                msg.attach(MIMEText(content, 'html'))
                
                server.send_message(msg)
                sent_count += 1
                
            except Exception as e:
                failed_emails.append({"email": user.email, "error": str(e)})
        
        # Log audit action for email sending
        admin = db.query(Admin).filter(Admin.email == current_user).first()
        user_role = admin.role if admin else "unknown"
        log_audit_action(db, current_user, user_role, "send_email", "email_campaign", template.id, f"Sent email '{template.name}' to {sent_count} students")
        
        
    except smtplib.SMTPAuthenticationError:
        raise HTTPException(status_code=400, detail="SMTP Authentication failed. Please check your email settings.")
    except smtplib.SMTPConnectError:
        raise HTTPException(status_code=400, detail="Cannot connect to SMTP server. Please check server address and port.")
    except Exception as e:
        error_msg = str(e)
        if "getaddrinfo failed" in error_msg:
            raise HTTPException(status_code=400, detail="Cannot resolve SMTP server address. Please check the server name.")
        else:
            raise HTTPException(status_code=500, detail=f"SMTP connection failed: {error_msg}")
    finally:
        if server:
            try:
                server.quit()
            except:
                pass
    
    return {
        "message": f"Email sent to {sent_count} students",
        "sent_count": sent_count,
        "failed_count": len(failed_emails),
        "failed_emails": failed_emails
    }



# Dashboard statistics endpoints
@app.get("/api/dashboard/stats")
def get_dashboard_stats(event_id: Optional[int] = None, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    from datetime import datetime, timedelta
    from sqlalchemy import func, and_
    
    # Base query with optional event filter
    base_query = db.query(User)
    if event_id:
        base_query = base_query.filter(User.eventId == event_id)
    
    # Total registrations from users table
    total_registrations = base_query.count() or 0
    
    # Today's registrations
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    today_registrations = base_query.filter(
        User.created_at >= today
    ).count() or 0
    
    # This week's registrations
    week_start = today - timedelta(days=today.weekday())
    week_registrations = base_query.filter(
        User.created_at >= week_start
    ).count() or 0
    
    # Get events count
    total_events = db.query(func.count(Event.id)).scalar() or 0
    
    return {
        "total_registrations": total_registrations,
        "today_registrations": today_registrations,
        "week_registrations": week_registrations,
        "total_events": total_events
    }

@app.get("/api/dashboard/registration-chart")
def get_registration_chart_data(event_id: Optional[int] = None, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    from datetime import datetime, timedelta
    from sqlalchemy import func
    
    # Get last 7 days registration data
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    seven_days_ago = today - timedelta(days=6)
    
    query = db.query(
        func.date(User.created_at).label('date'),
        func.count(User.id).label('count')
    ).filter(
        User.created_at >= seven_days_ago
    )
    
    if event_id:
        query = query.filter(User.eventId == event_id)
    
    daily_data = query.group_by(func.date(User.created_at)).order_by(func.date(User.created_at)).all()
    
    # Create labels and data for last 7 days
    labels = []
    data = []
    
    for i in range(7):
        date = seven_days_ago + timedelta(days=i)
        day_label = date.strftime('%a')  # Mon, Tue, etc.
        labels.append(day_label)
        
        # Find count for this date
        count = 0
        for daily in daily_data:
            if str(daily.date) == str(date.date()):
                count = daily.count
                break
        data.append(count)
    
    return {
        "labels": labels,
        "data": data
    }

@app.get("/api/dashboard/monthly-registrations")
def get_monthly_registrations(event_id: Optional[int] = None, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    from datetime import datetime, timedelta
    from sqlalchemy import func, extract
    
    # Get last 9 months data
    current_date = datetime.now()
    nine_months_ago = current_date - timedelta(days=270)
    
    query = db.query(
        extract('year', User.created_at).label('year'),
        extract('month', User.created_at).label('month'),
        func.count(User.id).label('count')
    ).filter(
        User.created_at >= nine_months_ago
    )
    
    if event_id:
        query = query.filter(User.eventId == event_id)
    
    monthly_data = query.group_by(
        extract('year', User.created_at),
        extract('month', User.created_at)
    ).order_by(
        extract('year', User.created_at),
        extract('month', User.created_at)
    ).all()
    
    # Create month labels and data
    month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    
    labels = []
    data = []
    
    # Generate last 9 months
    for i in range(9):
        target_date = current_date - timedelta(days=30 * (8 - i))
        month_label = month_names[target_date.month - 1]
        labels.append(month_label)
        
        # Find count for this month
        count = 0
        for monthly in monthly_data:
            if monthly.year == target_date.year and monthly.month == target_date.month:
                count = monthly.count
                break
        data.append(count)
    
    return {
        "labels": labels,
        "data": data
    }

@app.get("/api/dashboard/event-payments")
def get_event_payments(current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    from sqlalchemy import func
    
    try:
        # Get payment counts by event through user emails
        event_payments = db.query(
            Event.name,
            func.count(Payment.payment_id).label('paid_count')
        ).outerjoin(
            User, User.eventId == Event.id
        ).outerjoin(
            Payment, Payment.user_email == User.email
        ).filter(
            Payment.payment_status == "completed"
        ).group_by(Event.id, Event.name).order_by(Event.created_at).all()
        
        return {
            "labels": [event.name for event in event_payments],
            "data": [event.paid_count for event in event_payments]
        }
    except Exception as e:
        print(f"Error in get_event_payments: {str(e)}")
        # Return empty data if there's an error
        return {
            "labels": [],
            "data": []
        }

@app.get("/api/dashboard/attendance-stats")
def get_attendance_stats(event_id: Optional[int] = None, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    from sqlalchemy import func
    
    # Get attendance forms with optional event filter
    forms_query = db.query(Form).filter(Form.type == "attendance")
    if event_id:
        forms_query = forms_query.filter(Form.event_id == event_id)
    attendance_forms = forms_query.all()
    
    if not attendance_forms:
        return {
            "labels": ["No Data"],
            "data": [0]
        }
    
    # Get users with optional event filter
    users_query = db.query(User)
    if event_id:
        users_query = users_query.filter(User.eventId == event_id)
    users = users_query.all()
    attendance_counts = {"attended": 0, "partially_attended": 0, "not_attended": 0}
    
    for user in users:
        user_responses = 0
        for form in attendance_forms:
            response = db.query(FormResponse).filter(
                FormResponse.form_id == form.id,
                FormResponse.user_email == user.email
            ).first()
            if response:
                user_responses += 1
        
        if user_responses == len(attendance_forms) and len(attendance_forms) > 0:
            attendance_counts["attended"] += 1
        elif user_responses > 0:
            attendance_counts["partially_attended"] += 1
        else:
            attendance_counts["not_attended"] += 1
    
    return {
        "labels": ["Attended", "Partially Attended", "Not Attended"],
        "data": [attendance_counts["attended"], attendance_counts["partially_attended"], attendance_counts["not_attended"]]
    }

# Q/A System Endpoints - Remove duplicate and fix routing

@app.get("/api/qa/manager-questions/{event_id}")
def get_manager_questions(event_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Check if user is manager or admin
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not admin or admin.role not in ["manager", "admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    questions = db.query(QAQuestion).filter(
        QAQuestion.event_id == event_id,
        QAQuestion.status == "pending"
    ).order_by(QAQuestion.created_at.desc()).all()
    
    return [{
        "id": q.id,
        "user_name": q.user_name,
        "user_email": q.user_email,
        "question": q.question,
        "created_at": str(q.created_at)
    } for q in questions]

@app.post("/api/qa/toggle-event")
def toggle_qa_event(toggle_data: QAToggleRequest, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    try:
        print(f"Toggle QA request: event_id={toggle_data.event_id}, active={toggle_data.active}, user={current_user}")
        
        # Check if user is admin
        admin = db.query(Admin).filter(Admin.email == current_user).first()
        if not admin:
            print(f"Admin not found for email: {current_user}")
            raise HTTPException(status_code=403, detail="User not found")
        
        if admin.role != "admin":
            print(f"User {current_user} has role {admin.role}, not admin")
            raise HTTPException(status_code=403, detail="Only admins can toggle Q/A events")
        
        # Find the event first
        event = db.query(Event).filter(Event.id == toggle_data.event_id).first()
        if not event:
            print(f"Event not found with ID: {toggle_data.event_id}")
            raise HTTPException(status_code=404, detail=f"Event with ID {toggle_data.event_id} not found")
        
        print(f"Found event: {event.name} (ID: {event.id})")
        
        if toggle_data.active:
            # Turn off all other events first
            db.query(Event).update({Event.qa_active: 0})
            # Turn on the selected event
            event.qa_active = 1
            print(f"Enabled QA for event: {event.name}")
        else:
            # Turn off the selected event
            event.qa_active = 0
            print(f"Disabled QA for event: {event.name}")
        
        db.commit()
        print(f"Database committed successfully")
        
        return {"message": f"Q/A {'enabled' if toggle_data.active else 'disabled'} for event {event.name}"}
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"Unexpected error in toggle_qa_event: {str(e)}")
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

@app.post("/api/qa/manager-approve/{question_id}")
def manager_approve_question(question_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Check if user is manager or admin
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not admin or admin.role not in ["manager", "admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    question = db.query(QAQuestion).filter(QAQuestion.id == question_id).first()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    question.status = "manager_approved"
    question.manager_approved_at = datetime.utcnow()
    
    # Update user question count
    user_count = db.query(UserQuestionCount).filter(
        UserQuestionCount.event_id == question.event_id,
        UserQuestionCount.user_email == question.user_email
    ).first()
    
    if user_count:
        user_count.approved_questions += 1
        user_count.updated_at = datetime.utcnow()
    else:
        user_count = UserQuestionCount(
            event_id=question.event_id,
            user_email=question.user_email,
            user_name=question.user_name,
            approved_questions=1
        )
        db.add(user_count)
    
    db.commit()
    return {"message": "Question approved successfully"}

@app.post("/api/qa/manager-reject/{question_id}")
def manager_reject_question(question_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Check if user is manager or admin
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not admin or admin.role not in ["manager", "admin"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    question = db.query(QAQuestion).filter(QAQuestion.id == question_id).first()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    question.status = "rejected"
    question.admin_action = "rejected"
    question.admin_action_at = datetime.utcnow()
    
    db.commit()
    return {"message": "Question rejected successfully"}

@app.get("/api/qa/admin-questions/{event_id}")
def get_admin_questions(event_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Check if user is admin or presenter
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not admin or admin.role not in ["admin", "presenter"]:
        raise HTTPException(status_code=403, detail="Access denied")
    
    questions = db.query(QAQuestion).filter(
        QAQuestion.event_id == event_id,
        QAQuestion.status == "manager_approved"
    ).order_by(QAQuestion.manager_approved_at.desc()).all()
    
    return [{
        "id": q.id,
        "user_name": q.user_name,
        "user_email": q.user_email,
        "question": q.question,
        "created_at": str(q.created_at),
        "manager_approved_at": str(q.manager_approved_at)
    } for q in questions]

@app.post("/api/qa/admin-action")
def admin_action_question(action_data: QAActionRequest, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Check if user is admin
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not admin or admin.role != "admin":
        raise HTTPException(status_code=403, detail="Access denied")
    
    question = db.query(QAQuestion).filter(QAQuestion.id == action_data.question_id).first()
    if not question:
        raise HTTPException(status_code=404, detail="Question not found")
    
    question.admin_action = action_data.action
    question.admin_response = action_data.response if action_data.response else ""
    question.admin_action_at = datetime.utcnow()
    question.status = action_data.action
    
    db.commit()
    return {"message": f"Question {action_data.action} successfully"}

@app.get("/api/qa/user-questions")
def get_user_questions(email: str, db: Session = Depends(get_db)):
    """Get questions submitted by a specific user"""
    try:
        print(f"Fetching questions for user: {email}")
        
        # Check if there's an active Q/A session
        active_event = db.query(Event).filter(Event.qa_active == 1).first()
        if not active_event:
            print("No active Q/A session")
            return []
        
        questions = db.query(QAQuestion).filter(
            QAQuestion.user_email == email,
            QAQuestion.event_id == active_event.id
        ).order_by(QAQuestion.created_at.desc()).all()
        
        print(f"Found {len(questions)} questions for user {email}")
        
        return [{
            "id": q.id,
            "question": q.question,
            "status": q.status,
            "created_at": str(q.created_at),
            "admin_response": getattr(q, 'admin_response', None) or ""
        } for q in questions]
        
    except Exception as e:
        print(f"Error fetching user questions: {str(e)}")
        raise HTTPException(status_code=500, detail="Failed to fetch questions")

@app.get("/api/qa/top-students/{event_id}")
def get_top_students(event_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Get top students with college information
    
    # Join UserQuestionCount with User table to get college information
    top_students_with_college = db.query(
        UserQuestionCount.user_name,
        UserQuestionCount.user_email,
        UserQuestionCount.approved_questions,
        User.college_name
    ).join(
        User, UserQuestionCount.user_email == User.email
    ).filter(
        UserQuestionCount.event_id == event_id
    ).order_by(UserQuestionCount.approved_questions.desc()).limit(10).all()
    
    return [{
        "user_name": s.user_name,
        "user_email": s.user_email,
        "college_name": s.college_name or "N/A",
        "approved_questions": s.approved_questions
    } for s in top_students_with_college]

@app.get("/api/dashboard/colleges")
def get_colleges_stats(event_id: Optional[int] = None, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    from sqlalchemy import func
    
    # Base query with optional event filter
    base_query = db.query(User).filter(
        User.college_name.isnot(None),
        User.college_name != ''
    )
    if event_id:
        base_query = base_query.filter(User.eventId == event_id)
    
    # Get college statistics from users table
    college_stats = base_query.with_entities(
        User.college_name,
        func.count(User.id).label('count')
    ).group_by(User.college_name).order_by(func.count(User.id).desc()).limit(5).all()
    
    # Get total unique colleges count
    total_colleges = base_query.with_entities(
        func.count(func.distinct(User.college_name))
    ).scalar() or 0
    
    return {
        "colleges": [{
            "name": college.college_name,
            "count": college.count
        } for college in college_stats],
        "total_colleges": total_colleges
    }

# Missing endpoints for reports functionality
@app.get("/api/events/{event_id}/report-stats")
def get_event_report_stats(event_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    """Get basic stats for event reports page"""
    from sqlalchemy import func
    
    try:
        # Get total participants for this event
        total_participants = db.query(func.count(User.id)).filter(User.eventId == event_id).scalar() or 0
        
        # Calculate attendance rate based on attendance forms
        attendance_forms = db.query(Form).filter(
            Form.event_id == event_id,
            Form.type == "attendance"
        ).all()
    
        attended_count = 0
        if attendance_forms and total_participants > 0:
            users = db.query(User).filter(User.eventId == event_id).all()
            for user in users:
                user_responses = 0
                for form in attendance_forms:
                    response = db.query(FormResponse).filter(
                        FormResponse.form_id == form.id,
                        FormResponse.user_email == user.email
                    ).first()
                    if response:
                        user_responses += 1
                
                # Consider fully attended if responded to all forms
                if user_responses == len(attendance_forms):
                    attended_count += 1
        
        attendance_rate = round((attended_count / total_participants) * 100) if total_participants > 0 else 0
        
        # Calculate paid participants from payments table through user emails
        paid_count = db.query(func.count(Payment.payment_id)).join(
            User, User.email == Payment.user_email
        ).filter(
            User.eventId == event_id,
            Payment.payment_status == "completed"
        ).scalar() or 0
    
        return {
            "totalParticipants": total_participants,
            "attendanceRate": attendance_rate,
            "totalPaid": paid_count
        }
    except Exception as e:
        print(f"Error in get_event_report_stats: {str(e)}")
        return {
            "totalParticipants": 0,
            "attendanceRate": 0,
            "totalPaid": 0
        }

@app.get("/api/events/{event_id}/participants")
def get_event_participants(event_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    """Get participants for event detail report"""
    try:
        # Get all users for this event
        users = db.query(User).filter(User.eventId == event_id).all()
        
        # Get attendance forms for this event
        attendance_forms = db.query(Form).filter(
            Form.event_id == event_id,
            Form.type == "attendance"
        ).all()
        
        participants = []
        for user in users:
            # Calculate attendance status
            user_responses = 0
            for form in attendance_forms:
                response = db.query(FormResponse).filter(
                    FormResponse.form_id == form.id,
                    FormResponse.user_email == user.email
                ).first()
                if response:
                    user_responses += 1
            
            # Determine attendance status
            if user_responses == len(attendance_forms) and len(attendance_forms) > 0:
                attendance_status = "Attended"
            elif user_responses > 0:
                attendance_status = "Partially Attended"
            else:
                attendance_status = "Not Attended"
            
            # Check payment status from payments table
            payment = db.query(Payment).filter(
                Payment.user_email == user.email,
                Payment.payment_status == "completed"
            ).first()
            payment_status = "Paid" if payment else "Pending"
            
            # Get chat count for this user
            chat_count = db.query(QAQuestion).filter(
                QAQuestion.user_email == user.email,
                QAQuestion.event_id == event_id
            ).count()
            
            participants.append({
                "id": user.id,
                "name": user.name or f"{user.first_name or ''} {user.last_name or ''}".strip(),
                "email": user.email,
                "college": user.college_name or "",
                "phone": user.phone_number or "",
                "payment_status": payment_status,
                "attendance_status": attendance_status,
                "chat_count": chat_count,
                "created_at": user.created_at
            })
    
        return participants
    except Exception as e:
        print(f"Error in get_event_participants: {str(e)}")
        return []

@app.get("/api/events/{event_id}/analytics")
def get_event_analytics(event_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    from sqlalchemy import func, extract
    from datetime import datetime, timedelta
    
    try:
        # Get event details
        event = db.query(Event).filter(Event.id == event_id).first()
        if not event:
            return {
                "event_name": f"Event {event_id}",
                "total_registrations": 0,
                "colleges": [],
                "daily_registrations": [],
                "gender_distribution": [],
                "user_type_distribution": [],
                "attendance_stats": {"attended": 0, "partially_attended": 0, "not_attended": 0},
                "attendanceData": [],
                "paymentData": [],
                "collegeData": []
            }
        
        # Total registrations for this event
        total_registrations = db.query(func.count(User.id)).filter(User.eventId == event_id).scalar() or 0
        
        # Get all attendance forms for this event
        attendance_forms = db.query(Form).filter(
            Form.event_id == event_id,
            Form.type == "attendance"
        ).all()
        
        # Calculate attendance status for each user
        users = db.query(User).filter(User.eventId == event_id).all()
        attendance_counts = {"attended": 0, "partially_attended": 0, "not_attended": 0}
        
        for user in users:
            user_responses = 0
            for form in attendance_forms:
                response = db.query(FormResponse).filter(
                    FormResponse.form_id == form.id,
                    FormResponse.user_email == user.email
                ).first()
                if response:
                    user_responses += 1
            
            if user_responses == len(attendance_forms) and len(attendance_forms) > 0:
                attendance_counts["attended"] += 1
            elif user_responses > 0:
                attendance_counts["partially_attended"] += 1
            else:
                attendance_counts["not_attended"] += 1
        
        # College statistics with attendance
        college_stats = db.query(
            User.college_name,
            func.count(User.id).label('count')
        ).filter(
            User.eventId == event_id,
            User.college_name.isnot(None),
            User.college_name != ''
        ).group_by(User.college_name).order_by(func.count(User.id).desc()).all()
        
        college_stats_dict = {}
        for stat in college_stats:
            college_users = db.query(User).filter(
                User.eventId == event_id,
                User.college_name == stat.college_name
            ).all()
            
            college_attendance = {"attended": 0, "partially_attended": 0, "not_attended": 0}
            form_responses = {}
            
            # Initialize form response counts
            for form in attendance_forms:
                form_responses[f"{form.title}_responses"] = 0
            
            for user in college_users:
                user_responses = 0
                for form in attendance_forms:
                    response = db.query(FormResponse).filter(
                        FormResponse.form_id == form.id,
                        FormResponse.user_email == user.email
                    ).first()
                    if response:
                        user_responses += 1
                        form_responses[f"{form.title}_responses"] += 1
                
                if user_responses == len(attendance_forms) and len(attendance_forms) > 0:
                    college_attendance["attended"] += 1
                elif user_responses > 0:
                    college_attendance["partially_attended"] += 1
                else:
                    college_attendance["not_attended"] += 1
            
            college_stats_dict[stat.college_name] = {
                "registered": stat.count,
                **college_attendance,
                **form_responses
            }
        
        # Daily registration trend (last 30 days)
        thirty_days_ago = datetime.now() - timedelta(days=30)
        daily_registrations = db.query(
            func.date(User.created_at).label('date'),
            func.count(User.id).label('count')
        ).filter(
            User.eventId == event_id,
            User.created_at >= thirty_days_ago
        ).group_by(func.date(User.created_at)).order_by(func.date(User.created_at)).all()
        
        # Other stats
        gender_stats = db.query(
            User.gender,
            func.count(User.id).label('count')
        ).filter(
            User.eventId == event_id,
            User.gender.isnot(None),
            User.gender != ''
        ).group_by(User.gender).all()
        
        user_type_stats = db.query(
            User.user_type,
            func.count(User.id).label('count')
        ).filter(
            User.eventId == event_id,
            User.user_type.isnot(None),
            User.user_type != ''
        ).group_by(User.user_type).all()
        
        utm_stats = db.query(
            User.utm_source,
            func.count(User.id).label('count')
        ).filter(
            User.eventId == event_id,
            User.utm_source.isnot(None),
            User.utm_source != ''
        ).group_by(User.utm_source).all()
        
        # Calculate attendance data for charts
        attended = 0
        partially_attended = 0
        not_attended = 0
        
        for user in users:
            user_responses = 0
            for form in attendance_forms:
                response = db.query(FormResponse).filter(
                    FormResponse.form_id == form.id,
                    FormResponse.user_email == user.email
                ).first()
                if response:
                    user_responses += 1
            
            if user_responses == len(attendance_forms) and len(attendance_forms) > 0:
                attended += 1
            elif user_responses > 0:
                partially_attended += 1
            else:
                not_attended += 1
        
        # Payment data from actual payments table
        paid_count = db.query(func.count(Payment.payment_id)).join(
            User, User.email == Payment.user_email
        ).filter(
            User.eventId == event_id,
            Payment.payment_status == "completed"
        ).scalar() or 0
        
        payment_data = [
            {"name": "Paid", "value": paid_count},
            {"name": "Pending", "value": total_registrations - paid_count}
        ]
        
        # College data for charts
        college_data = [{"name": stat.college_name, "count": stat.count} for stat in college_stats]
        
        attendance_data = [
            {"name": "Attended", "value": attended},
            {"name": "Partially Attended", "value": partially_attended},
            {"name": "Not Attended", "value": not_attended}
        ]
    
        return {
            "event_name": event.name,
            "total_registrations": total_registrations,
            "colleges": [{
                "name": stat.college_name,
                "count": stat.count
            } for stat in college_stats],
            "college_stats": college_stats_dict,
            "daily_registrations": [{
                "date": str(stat.date),
                "count": stat.count
            } for stat in daily_registrations],
            "gender_distribution": [{
                "gender": stat.gender,
                "count": stat.count
            } for stat in gender_stats],
            "user_type_distribution": [{
                "type": stat.user_type,
                "count": stat.count
            } for stat in user_type_stats],
            "utm_sources": [{
                "source": stat.utm_source or "Unknown",
                "count": stat.count
            } for stat in utm_stats],
            "attendance_stats": attendance_counts,
            "attendanceData": attendance_data,
            "paymentData": payment_data,
            "collegeData": college_data
        }
    except Exception as e:
        print(f"Error in get_event_analytics: {str(e)}")
        return {
            "event_name": f"Event {event_id}",
            "total_registrations": 0,
            "colleges": [],
            "daily_registrations": [],
            "gender_distribution": [],
            "user_type_distribution": [],
            "attendance_stats": {"attended": 0, "partially_attended": 0, "not_attended": 0},
            "attendanceData": [],
            "paymentData": [],
            "collegeData": []
        }

@app.get("/api/events/{event_id}/utm-sources")
def get_utm_sources(event_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    from sqlalchemy import func
    
    utm_stats = db.query(
        User.utm_source,
        func.count(User.id).label('count')
    ).filter(
        User.eventId == event_id,
        User.utm_source.isnot(None),
        User.utm_source != ''
    ).group_by(User.utm_source).order_by(func.count(User.id).desc()).all()
    
    return [{
        "source": stat.utm_source,
        "count": stat.count
    } for stat in utm_stats]

@app.get("/api/events/{event_id}/attendance-forms")
def get_event_attendance_forms(event_id: int, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    forms = db.query(Form).filter(
        Form.event_id == event_id,
        Form.type == "attendance"
    ).all()
    
    return [{
        "id": form.id,
        "title": form.title,
        "description": form.description,
        "created_at": str(form.created_at),
        "is_active": form.is_active
    } for form in forms]

# Audit logging helper function
def log_audit_action(db: Session, user_email: str, user_role: str, action: str, resource_type: str, resource_id: str = None, details: str = None, ip_address: str = None):
    from database import AuditLog, Admin
    import json
    
    # Get user name
    admin = db.query(Admin).filter(Admin.email == user_email).first()
    user_name = admin.email if admin else user_email
    
    audit_log = AuditLog(
        user_email=user_email,
        user_name=user_name,
        user_role=user_role,
        action=action,
        resource_type=resource_type,
        resource_id=str(resource_id) if resource_id else None,
        details=details,
        ip_address=ip_address
    )
    db.add(audit_log)
    db.commit()

@app.get("/api/logs")
def get_audit_logs(user_filter: str = None, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    """Get audit logs (admin only)"""
    from database import Admin, AuditLog
    
    # Check if current user is admin
    current_admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not current_admin or current_admin.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view logs")
    
    # Build query with optional user filter
    query = db.query(AuditLog)
    if user_filter:
        query = query.filter(AuditLog.user_email.contains(user_filter))
    
    logs = query.order_by(AuditLog.created_at.desc()).limit(1000).all()
    
    return [{
        "id": log.id,
        "user_email": log.user_email,
        "user_name": log.user_name,
        "user_role": log.user_role,
        "action": log.action,
        "resource_type": log.resource_type,
        "resource_id": log.resource_id,
        "details": log.details,
        "ip_address": log.ip_address,
        "created_at": str(log.created_at)
    } for log in logs]

# User Management Endpoints
class UserUpdate(BaseModel):
    name: Optional[str] = None
    email: Optional[str] = None
    role: Optional[str] = None
    college: Optional[str] = None
    attendance_status: Optional[str] = None

@app.get("/api/users")
def get_users(
    current_user: str = Depends(verify_token),
    db: Session = Depends(get_db),
    college: Optional[str] = None,
    status: Optional[str] = None,
    attendance: Optional[str] = None,
    event: Optional[str] = None,
    search: Optional[str] = None
):
    # Get users with event names using join
    query = db.query(User, Event.name.label('event_name')).outerjoin(
        Event, User.eventId == Event.id
    )
    
    if college:
        query = query.filter(User.college_name == college)
    if event:
        event_obj = db.query(Event).filter(Event.name == event).first()
        if event_obj:
            query = query.filter(User.eventId == event_obj.id)
    if search:
        query = query.filter(
            User.name.contains(search) | 
            User.email.contains(search) |
            User.college_name.contains(search)
        )
    
    users_with_events = query.all()
    users = [{
        "id": str(user.User.id),
        "name": user.User.name or f"{user.User.first_name or ''} {user.User.last_name or ''}".strip(),
        "email": user.User.email,
        "college": user.User.college_name or "",
        "phone": user.User.phone_number or "",
        "event_id": user.User.eventId,
        "event": user.event_name or "Unknown Event",
        "registered_at": str(user.User.created_at),
        "created_date": str(user.User.created_at)
    } for user in users_with_events]
    total_count = len(users)
    
    # Log audit action
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    user_role = admin.role if admin else "unknown"
    log_audit_action(db, current_user, user_role, "view_users", "user_management", None, f"Viewed users list with filters")
    
    return {"users": users, "total_count": total_count}

@app.get("/api/users/report")
def generate_users_report(
    current_user: str = Depends(verify_token),
    db: Session = Depends(get_db),
    college: Optional[str] = None,
    event: Optional[str] = None,
    format: str = "excel"
):
    from fastapi.responses import StreamingResponse
    import io
    from datetime import datetime
    
    try:
        import pandas as pd
    except ImportError:
        raise HTTPException(status_code=500, detail="pandas is required for report generation")
    
    # Company configuration
    COMPANY_CONFIG = {
        "name": "Kambaa Incorporation",
        "address": "10th Floor, North Wing, Pricol Caledon Square\nAvinashi Road, Coimbatore",
        "contact": "Email: contact@kambaa.in | Phone: +91 7094490097",
        "logo_path": r"C:\Users\DELL\Downloads\Event_Dashboard\Event_Dashboard\frontend\Kambaa Logo.png"
    }
    
    # Get users data with proper attendance calculation
    users_data = []
    
    # Get all users with filters using direct database queries
    query = db.query(User)
    
    if college:
        query = query.filter(User.college_name == college)
    if event:
        event_obj = db.query(Event).filter(Event.name == event).first()
        if event_obj:
            query = query.filter(User.eventId == event_obj.id)
    
    users_db = query.all()
    users = [{
        "name": u.name or f"{u.first_name or ''} {u.last_name or ''}".strip(),
        "email": u.email,
        "college": u.college_name or "",
        "phone": u.phone_number or "",
        "event_id": u.eventId,
        "registered_at": str(u.created_at)
    } for u in users_db]
    
    # Calculate proper attendance for each user
    for user_data in users:
        user_email = user_data['email']
        event_id = None
        
        # Get event ID if event filter is applied
        if event:
            event_obj = db.query(Event).filter(Event.name == event).first()
            if event_obj:
                event_id = event_obj.id
        
        # Calculate attendance based on form responses
        attendance_status = "Not Attended"
        if event_id:
            # Get attendance forms for this event
            attendance_forms = db.query(Form).filter(
                Form.event_id == event_id,
                Form.type == "attendance"
            ).all()
            
            if attendance_forms:
                user_responses = 0
                for form in attendance_forms:
                    response = db.query(FormResponse).filter(
                        FormResponse.form_id == form.id,
                        FormResponse.user_email == user_email
                    ).first()
                    if response:
                        user_responses += 1
                
                if user_responses == len(attendance_forms):
                    attendance_status = "Attended"
                elif user_responses > 0:
                    attendance_status = "Partially Attended"
        
        # Calculate chat count (QA questions) for this user
        chat_count = 0
        if event_id:
            chat_count = db.query(QAQuestion).filter(
                QAQuestion.user_email == user_email,
                QAQuestion.event_id == event_id
            ).count()
        
        users_data.append({
            'Name': user_data['name'],
            'Email': user_data['email'],
            'College': user_data['college'],
            'Attendance': attendance_status,
            'Chat Count': chat_count
        })
    
    # Get event details for report header
    event_name = event or "All Events"
    event_date = ""
    if event:
        event_obj = db.query(Event).filter(Event.name == event).first()
        if event_obj:
            event_date = str(event_obj.event_date)
    
    if format.lower() == "pdf":
        try:
            from reportlab.lib.pagesizes import A4, letter
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            import os
            
            # Generate PDF report
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
            elements = []
            
            # Styles
            styles = getSampleStyleSheet()
            
            # Company logo (if exists)
            if os.path.exists(COMPANY_CONFIG["logo_path"]):
                try:
                    logo = Image(COMPANY_CONFIG["logo_path"], width=2*inch, height=0.8*inch)
                    elements.append(logo)
                    elements.append(Spacer(1, 12))
                except:
                    pass
            
            # Company header
            company_style = ParagraphStyle(
                'CompanyHeader',
                parent=styles['Normal'],
                fontSize=14,
                fontName='Helvetica-Bold',
                alignment=1,
                spaceAfter=6
            )
            elements.append(Paragraph(COMPANY_CONFIG["name"], company_style))
            
            address_style = ParagraphStyle(
                'Address',
                parent=styles['Normal'],
                fontSize=10,
                alignment=1,
                spaceAfter=3
            )
            elements.append(Paragraph(COMPANY_CONFIG["address"], address_style))
            elements.append(Paragraph(COMPANY_CONFIG["contact"], address_style))
            elements.append(Spacer(1, 20))
            
            # Report title
            title_style = ParagraphStyle(
                'ReportTitle',
                parent=styles['Heading1'],
                fontSize=16,
                fontName='Helvetica-Bold',
                alignment=1,
                spaceAfter=10
            )
            elements.append(Paragraph("Event Registration Report", title_style))
            
            # Event details
            if event_name != "All Events":
                event_style = ParagraphStyle(
                    'EventDetails',
                    parent=styles['Normal'],
                    fontSize=12,
                    fontName='Helvetica-Bold',
                    alignment=1,
                    spaceAfter=5
                )
                elements.append(Paragraph(f"Event: {event_name}", event_style))
                if event_date:
                    # Format the date nicely
                    try:
                        from datetime import datetime
                        parsed_date = datetime.strptime(event_date.split(' ')[0], '%Y-%m-%d')
                        formatted_date = parsed_date.strftime('%B %d, %Y')
                        elements.append(Paragraph(f"Date: {formatted_date}", event_style))
                    except:
                        elements.append(Paragraph(f"Date: {event_date}", event_style))
            else:
                # Show date range for all events if available
                elements.append(Paragraph("Report: All Events", event_style))
            
            elements.append(Spacer(1, 20))
            
            # Table data (5-column layout with chat count)
            if users_data:
                table_data = [['Name', 'Email', 'College', 'Attendance', 'Chats']]
                
                for user in users_data:
                    # Handle long college names by wrapping to next line
                    college_name = user['College']
                    if len(college_name) > 35:
                        # Split long college names into multiple lines
                        words = college_name.split(' ')
                        lines = []
                        current_line = ''
                        for word in words:
                            if len(current_line + ' ' + word) <= 35:
                                current_line += (' ' + word) if current_line else word
                            else:
                                if current_line:
                                    lines.append(current_line)
                                current_line = word
                        if current_line:
                            lines.append(current_line)
                        college_name = '\n'.join(lines)
                    
                    table_data.append([
                        user['Name'],
                        user['Email'],
                        college_name,
                        user['Attendance'],
                        str(user['Chat Count'])
                    ])
                
                table = Table(table_data, colWidths=[1.5*inch, 2*inch, 2.2*inch, 1.3*inch, 1*inch])
                table.setStyle(TableStyle([
                    # Professional blue headers
                    ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.1, 0.45, 0.91)),  # Professional blue
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('VALIGN', (0, 0), (-1, -1), 'TOP'),  # Changed to TOP for better multi-line display
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, 0), 12),
                    # Data rows
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),  # Slightly smaller font for better fit
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.7, 0.7, 0.7)),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 1), (-1, -1), 10),  # More padding for multi-line content
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
                ]))
                elements.append(table)
                
                # Attendance summary
                elements.append(Spacer(1, 20))
                attended_count = sum(1 for user in users_data if user['Attendance'] == 'Attended')
                partial_count = sum(1 for user in users_data if user['Attendance'] == 'Partially Attended')
                not_attended_count = sum(1 for user in users_data if user['Attendance'] == 'Not Attended')
                
                summary_style = ParagraphStyle(
                    'Summary',
                    parent=styles['Normal'],
                    fontSize=11,
                    fontName='Helvetica-Bold',
                    spaceAfter=3
                )
                elements.append(Paragraph("Attendance Summary:", summary_style))
                elements.append(Paragraph(f"Total Registrations: {len(users_data)}", styles['Normal']))
                elements.append(Paragraph(f"Attended: {attended_count}", styles['Normal']))
                elements.append(Paragraph(f"Partially Attended: {partial_count}", styles['Normal']))
                elements.append(Paragraph(f"Not Attended: {not_attended_count}", styles['Normal']))
            else:
                elements.append(Paragraph("No data available for the selected filters.", styles['Normal']))
            
            # Footer
            elements.append(Spacer(1, 30))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                alignment=1,
                textColor=colors.Color(0.5, 0.5, 0.5)
            )
            elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
            elements.append(Paragraph(COMPANY_CONFIG["contact"], footer_style))
            
            doc.build(elements)
            buffer.seek(0)
            
            # Dynamic filename
            filename = f"Kambaa_Event_Report_{event_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            
            return StreamingResponse(
                io.BytesIO(buffer.read()),
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
        except ImportError:
            raise HTTPException(status_code=400, detail="PDF generation requires reportlab. Please use Excel format instead.")
    
    else:
        # Generate Excel report with professional formatting
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.drawing.image import Image as OpenpyxlImage
            import os
            
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Event Registration Report"
            
            # Professional blue color scheme
            blue_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
            white_font = Font(color="FFFFFF", bold=True, size=11)
            header_font = Font(bold=True, size=12)
            
            # Start without logo for Excel reports
            row_offset = 1
            
            # Company header
            ws.merge_cells(f'A{row_offset}:D{row_offset}')
            ws[f'A{row_offset}'] = COMPANY_CONFIG["name"]
            ws[f'A{row_offset}'].font = Font(bold=True, size=16)
            ws[f'A{row_offset}'].alignment = Alignment(horizontal='center')
            
            row_offset += 1
            ws.merge_cells(f'A{row_offset}:D{row_offset}')
            ws[f'A{row_offset}'] = COMPANY_CONFIG["address"]
            ws[f'A{row_offset}'].alignment = Alignment(horizontal='center')
            
            row_offset += 1
            ws.merge_cells(f'A{row_offset}:D{row_offset}')
            ws[f'A{row_offset}'] = COMPANY_CONFIG["contact"]
            ws[f'A{row_offset}'].alignment = Alignment(horizontal='center')
            
            row_offset += 3
            
            # Report title
            ws.merge_cells(f'A{row_offset}:D{row_offset}')
            ws[f'A{row_offset}'] = "Event Registration Report"
            ws[f'A{row_offset}'].font = header_font
            ws[f'A{row_offset}'].alignment = Alignment(horizontal='center')
            
            row_offset += 1
            
            # Event details
            if event_name != "All Events":
                ws.merge_cells(f'A{row_offset}:D{row_offset}')
                ws[f'A{row_offset}'] = f"Event: {event_name}"
                ws[f'A{row_offset}'].font = Font(bold=True)
                ws[f'A{row_offset}'].alignment = Alignment(horizontal='center')
                row_offset += 1
                
                if event_date:
                    ws.merge_cells(f'A{row_offset}:D{row_offset}')
                    # Format the date nicely for Excel
                    try:
                        from datetime import datetime
                        parsed_date = datetime.strptime(event_date.split(' ')[0], '%Y-%m-%d')
                        formatted_date = parsed_date.strftime('%B %d, %Y')
                        ws[f'A{row_offset}'] = f"Date: {formatted_date}"
                    except:
                        ws[f'A{row_offset}'] = f"Date: {event_date}"
                    ws[f'A{row_offset}'].alignment = Alignment(horizontal='center')
                    row_offset += 1
            else:
                ws.merge_cells(f'A{row_offset}:D{row_offset}')
                ws[f'A{row_offset}'] = "Report: All Events"
                ws[f'A{row_offset}'].font = Font(bold=True)
                ws[f'A{row_offset}'].alignment = Alignment(horizontal='center')
                row_offset += 1
            
            if college:
                ws.merge_cells(f'A{row_offset}:D{row_offset}')
                ws[f'A{row_offset}'] = f"College: {college}"
                ws[f'A{row_offset}'].font = Font(bold=True)
                ws[f'A{row_offset}'].alignment = Alignment(horizontal='center')
                row_offset += 1
            
            row_offset += 2
            
            # Headers with professional blue styling
            headers = ['Name', 'Email', 'College', 'Attendance', 'Chats']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row_offset, column=col, value=header)
                cell.fill = blue_fill
                cell.font = white_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                # Highlight the Chat Count column header
                if header == 'Chats':
                    cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")  # Orange highlight
            
            # Data rows
            for row_idx, user in enumerate(users_data, row_offset + 1):
                data = [user['Name'], user['Email'], user['College'], user['Attendance'], user['Chat Count']]
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row_idx, column=col, value=value)
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    cell.border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    # Highlight the Chat Count column data
                    if col == 5:  # Chat Count column
                        cell.fill = PatternFill(start_color="FFE5D9", end_color="FFE5D9", fill_type="solid")  # Light orange
            
            # Auto-adjust column widths - added chat count column
            column_widths = [25, 35, 35, 15, 12]  # Name, Email, College, Attendance, Chats
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + col)].width = width
            
            # Attendance summary at bottom
            summary_row = row_offset + len(users_data) + 3
            attended_count = sum(1 for user in users_data if user['Attendance'] == 'Attended')
            partial_count = sum(1 for user in users_data if user['Attendance'] == 'Partially Attended')
            not_attended_count = sum(1 for user in users_data if user['Attendance'] == 'Not Attended')
            
            ws[f'A{summary_row}'] = "Attendance Summary:"
            ws[f'A{summary_row}'].font = Font(bold=True)
            
            ws[f'A{summary_row + 1}'] = f"Total Registrations: {len(users_data)}"
            ws[f'A{summary_row + 1}'].font = Font(bold=True)
            
            ws[f'A{summary_row + 2}'] = f"Attended: {attended_count}"
            ws[f'A{summary_row + 3}'] = f"Partially Attended: {partial_count}"
            ws[f'A{summary_row + 4}'] = f"Not Attended: {not_attended_count}"
            
            # Save to buffer
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Dynamic filename
            filename = f"Kambaa_Event_Report_{event_name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                io.BytesIO(buffer.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )
            
        except ImportError:
            # Fallback to pandas if openpyxl is not available
            df = pd.DataFrame(users_data)
            
            buffer = io.BytesIO()
            with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Event Registration Report', index=False)
                
                workbook = writer.book
                worksheet = writer.sheets['Event Registration Report']
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
            buffer.seek(0)
            filename = f"Event_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            return StreamingResponse(
                io.BytesIO(buffer.read()),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"}
            )

@app.get("/api/users/{user_id}")
def get_user(user_id: str, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    user_db = db.query(User).filter(User.id == int(user_id)).first()
    if not user_db:
        raise HTTPException(status_code=404, detail="User not found")
    
    user = {
        "id": str(user_db.id),
        "name": user_db.name or f"{user_db.first_name or ''} {user_db.last_name or ''}".strip(),
        "email": user_db.email,
        "college": user_db.college_name or "",
        "phone": user_db.phone_number or "",
        "event_id": user_db.eventId,
        "registered_at": str(user_db.created_at)
    }
    
    # Log audit action
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    user_role = admin.role if admin else "unknown"
    log_audit_action(db, current_user, user_role, "view_user", "user", user_id, f"Viewed user: {user.get('name')}")
    
    return user

@app.put("/api/users/{user_id}")
def update_user(user_id: str, user_data: UserUpdate, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    update_data = {k: v for k, v in user_data.dict().items() if v is not None}
    if not update_data:
        raise HTTPException(status_code=400, detail="No data to update")
    
    user_db = db.query(User).filter(User.id == int(user_id)).first()
    if not user_db:
        raise HTTPException(status_code=404, detail="User not found")
    
    for key, value in update_data.items():
        if hasattr(user_db, key):
            setattr(user_db, key, value)
    
    db.commit()
    
    # Log audit action
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    user_role = admin.role if admin else "unknown"
    log_audit_action(db, current_user, user_role, "edit_user", "user", user_id, f"Updated user with data: {update_data}")
    
    return {"message": "User updated successfully"}

@app.delete("/api/users/{user_id}")
def delete_user(user_id: str, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    # Get user info before deletion for audit log
    user_db = db.query(User).filter(User.id == int(user_id)).first()
    if not user_db:
        raise HTTPException(status_code=404, detail="User not found")
    
    user_name = user_db.name or f"{user_db.first_name or ''} {user_db.last_name or ''}".strip()
    
    db.delete(user_db)
    db.commit()
    
    # Log audit action
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    user_role = admin.role if admin else "unknown"
    log_audit_action(db, current_user, user_role, "delete_user", "user", user_id, f"Deleted user: {user_name}")
    
    return {"message": "User deleted successfully"}

@app.post("/api/users/{user_id}/toggle-status")
def toggle_user_status(user_id: str, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    user_db = db.query(User).filter(User.id == int(user_id)).first()
    if not user_db:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Toggle status (assuming there's a status field)
    user_db.is_active = not getattr(user_db, 'is_active', True)
    db.commit()
    
    # Log audit action
    admin = db.query(Admin).filter(Admin.email == current_user).first()
    user_role = admin.role if admin else "unknown"
    log_audit_action(db, current_user, user_role, "toggle_user_status", "user", user_id, f"Toggled user status")
    
    return {"message": "User status toggled successfully"}



@app.get("/api/events/by-college")
def get_events_by_college(
    current_user: str = Depends(verify_token),
    college: Optional[str] = None,
    db: Session = Depends(get_db)
):
    query = db.query(Event)
    if college:
        # Get events that have users from this college
        query = query.join(User, User.eventId == Event.id).filter(User.college_name == college)
    
    events = query.distinct().all()
    return {"events": [event.name for event in events]}

@app.get("/api/admins")
def get_admins(current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    """Get list of all admins and managers (admin only)"""
    from database import Admin
    
    # Check if current user is admin
    current_admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not current_admin or current_admin.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can view admin list")
    
    # Get all admins, managers, and presenters
    admins = db.query(Admin).filter(Admin.role == "admin").all()
    managers = db.query(Admin).filter(Admin.role == "manager").all()
    presenters = db.query(Admin).filter(Admin.role == "presenter").all()
    
    return {
        "admins": [{
            "id": admin.id,
            "email": admin.email,
            "role": admin.role,
            "created_at": admin.created_at.isoformat() if admin.created_at else None
        } for admin in admins],
        "managers": [{
            "id": manager.id,
            "email": manager.email,
            "role": manager.role,
            "created_at": manager.created_at.isoformat() if manager.created_at else None
        } for manager in managers],
        "presenters": [{
            "id": presenter.id,
            "email": presenter.email,
            "role": presenter.role,
            "created_at": presenter.created_at.isoformat() if presenter.created_at else None
        } for presenter in presenters]
    }

class DeleteAdminRequest(BaseModel):
    admin_id: int

@app.post("/api/auth/delete-admin")
def delete_admin(delete_request: DeleteAdminRequest, current_user: str = Depends(verify_token), db: Session = Depends(get_db)):
    """Delete admin or manager (admin only)"""
    from database import Admin
    
    # Check if current user is admin
    current_admin = db.query(Admin).filter(Admin.email == current_user).first()
    if not current_admin or current_admin.role != "admin":
        raise HTTPException(status_code=403, detail="Only admins can delete users")
    
    # Find the admin to delete
    admin_to_delete = db.query(Admin).filter(Admin.id == delete_request.admin_id).first()
    if not admin_to_delete:
        raise HTTPException(status_code=404, detail="Admin not found")
    
    # Prevent self-deletion
    if admin_to_delete.email == current_user:
        raise HTTPException(status_code=400, detail="Cannot delete your own account")
    
    # Store info for audit log
    deleted_email = admin_to_delete.email
    deleted_role = admin_to_delete.role
    
    # Delete the admin
    db.delete(admin_to_delete)
    db.commit()
    
    # Log audit action
    log_audit_action(db, current_user, current_admin.role, "delete_admin", "admin", str(delete_request.admin_id), f"Deleted {deleted_role}: {deleted_email}")
    
    return {"message": f"{deleted_role.title()} deleted successfully"}

@app.websocket("/ws/forms/{form_id}")
async def websocket_endpoint(websocket: WebSocket, form_id: str):
    await manager.connect(websocket, form_id)
    try:
        while True:
            data = await websocket.receive_text()
            # Echo back for connection testing
            await manager.send_personal_message(f"Connected to form {form_id}", websocket)
    except WebSocketDisconnect:
        manager.disconnect(websocket, form_id)

@app.websocket("/ws/qa/{user_email}")
async def qa_websocket_endpoint(websocket: WebSocket, user_email: str):
    await qa_manager.connect(websocket, user_email)
    try:
        while True:
            data = await websocket.receive_text()
            # Handle incoming messages from user
            message_data = json.loads(data) if data else {}
            if message_data.get("type") == "join":
                await qa_manager.send_personal_message(
                    json.dumps({"type": "connected", "message": "Connected to Q/A session"}),
                    user_email
                )
    except WebSocketDisconnect:
        qa_manager.disconnect(user_email)
    except Exception as e:
        print(f"WebSocket error for user {user_email}: {str(e)}")
        qa_manager.disconnect(user_email)

@app.options("/{path:path}")
def options_handler(path: str):
    return {"message": "OK"}

@app.get("/")
def root():
    return {"message": "Dashboard API is running"}

@app.get("/health")
def health_check():
    return {"status": "healthy", "message": "API is working"}

@app.get("/api/health")
def api_health_check():
    return {"status": "healthy", "message": "API endpoints are working"}

@app.post("/api/qa/check-session")
def check_qa_session(validation_data: QAValidationRequest, db: Session = Depends(get_db)):
    """Check if user session is still valid"""
    try:
        # Get active event
        active_event = db.query(Event).filter(Event.qa_active == 1).first()
        if not active_event:
            return {"valid": False, "message": "No active Q/A session"}
        
        # Check if user exists and is valid
        user = db.query(User).filter(
            User.email == validation_data.email,
            User.registration_id == validation_data.registration_id
        ).first()
        
        if not user:
            return {"valid": False, "message": "Invalid user credentials"}
        
        user_name = user.name or f"{user.first_name or ''} {user.last_name or ''}".strip()
        
        return {
            "valid": True,
            "user_name": user_name,
            "event_name": active_event.name,
            "event_id": active_event.id
        }
        
    except Exception as e:
        print(f"Error checking session: {str(e)}")
        return {"valid": False, "message": "Session check failed"}

@app.get("/api/qa/debug")
def qa_debug_info(db: Session = Depends(get_db)):
    """Debug endpoint to check Q/A system status"""
    try:
        # Get all events with Q/A status
        events = db.query(Event).all()
        events_info = [{
            "id": e.id,
            "name": e.name,
            "qa_active": getattr(e, 'qa_active', 0)
        } for e in events]
        
        # Get active Q/A event
        active_event = db.query(Event).filter(Event.qa_active == 1).first()
        
        # Get sample users
        sample_users = db.query(User).limit(5).all()
        users_info = [{
            "email": u.email,
            "registration_id": u.registration_id,
            "name": u.name
        } for u in sample_users]
        
        # Get Q/A questions count
        questions_count = db.query(QAQuestion).count()
        
        return {
            "status": "ok",
            "events": events_info,
            "active_qa_event": {
                "id": active_event.id,
                "name": active_event.name,
                "qa_active": active_event.qa_active
            } if active_event else None,
            "sample_users": users_info,
            "total_questions": questions_count,
            "websocket_connections": len(qa_manager.active_connections),
            "message": "Q/A system is working properly" if active_event else "No active Q/A session"
        }
    except Exception as e:
        return {
            "status": "error",
            "error": str(e)
        }


# Email unsubscribe functionality using FastAPI
class UnsubscribeRequest(BaseModel):
    email: str
    reason: Optional[str] = None

def init_unsubscribe_db():
    import sqlite3
    conn = sqlite3.connect('students_db')
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS unsubscribed_emails (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            email TEXT UNIQUE NOT NULL,
            reason TEXT,
            unsubscribed_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()

@app.post("/api/unsubscribe")
def unsubscribe(request: UnsubscribeRequest):
    import sqlite3
    
    email = request.email.strip().lower()
    reason = request.reason.strip() if request.reason else ""
    
    if not email:
        raise HTTPException(status_code=400, detail="Email cannot be empty")
    
    try:
        init_unsubscribe_db()  # Ensure table exists
        conn = sqlite3.connect('email_management.db')
        cursor = conn.cursor()
        
        cursor.execute(
            'INSERT OR REPLACE INTO unsubscribed_emails (email, reason) VALUES (?, ?)',
            (email, reason)
        )
        
        conn.commit()
        conn.close()
        
        return {
            'message': 'Successfully unsubscribed',
            'email': email
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail="Database error occurred")

def is_email_unsubscribed(email):
    import sqlite3
    try:
        conn = sqlite3.connect('email_management.db')
        cursor = conn.cursor()
        cursor.execute('SELECT 1 FROM unsubscribed_emails WHERE email = ?', (email.lower(),))
        result = cursor.fetchone()
        conn.close()
        return result is not None
    except:
        return False

if __name__ == "__main__":
    print("Starting Event Dashboard API...")
    print("Q/A Debug endpoint: http://localhost:8000/api/qa/debug")
    print("User Chat Login: http://localhost:8000/user-chat-login")
    uvicorn.run(app, host="0.0.0.0", port=8000)